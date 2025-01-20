from django.core.management.base import BaseCommand, no_translations
from django.db import IntegrityError
from openpyxl import load_workbook

from financial_orgs.models import FinancialOrganization


class Command(BaseCommand):
    help = 'Import data from an Excel file into the FinancialOrganization model'

    def handle(self, *args, **kwargs):
        file_path = "finOrgs.xlsx" 

        if FinancialOrganization.objects.count() == 0:
            try:
                wb = load_workbook(file_path)
                sheet = wb.active

                for row in sheet.iter_rows(min_row=2, values_only=True):
                    #replace null with empty string
                    row = tuple("" if cell is None else cell for cell in row)
                    
                    FinancialOrganization.objects.create(
                        name=row[0],
                        first_director=row[1],
                        board_of_directors=row[2],
                        chairman_of_the_board=row[3],
                        board_members=row[4],
                        other_executives=row[5],
                        director=row[6],
                        chief_accountant=row[7],
                        bin=row[8],
                        address=row[9],
                        phone_number=row[10],
                        email=row[11],
                        website=row[12],
                        license=row[13],
                        branches=row[14],
                    )
                self.stdout.write(self.style.SUCCESS("Data inserted successfully into FinancialOrganization model."))
            except IntegrityError as e:
                self.stdout.write(self.style.ERROR(f"Error inserting data: {e}"))
            except Exception as e:
                self.stdout.write(self.style.ERROR(f"An error occurred: {e}"))
        else:
            self.stdout.write(self.style.WARNING("The FinancialOrganization table is not empty."))
